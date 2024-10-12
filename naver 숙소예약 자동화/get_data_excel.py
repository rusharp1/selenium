import openpyxl

# Excel 파일에서 호텔 검색 데이터 읽기
def search_get_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    ws = workbook.active
    
    excel_value = []
    for row in ws.iter_rows(min_row = 2, max_col=6, values_only=True):
        row_list = list(row)
        # 마지막 열의 데이터를 리스트로 변환
        row_list[-1] = [item.strip() for item in str(row[-1]).split(',')] if row[-1] else []
        excel_value.append(row_list)
    return excel_value

# Excel 파일에서 필터 데이터 읽기
def filter_get_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    ws = workbook.active

    excel_value = []
    for row in ws.iter_rows(min_row = 2, min_col=7,max_col=13, values_only=True):
        row_list = []
        for row_value in row:
            if row_value:
                row_list.append([item.strip() for item in str(row_value).split(',')])
            else:
                row_list.append([])
        excel_value.append(row_list)
    # 
    return excel_value

if __name__== "__main__":

    file_path = r"selenium\naver 숙소예약 자동화\data_value.xlsx"
    filter_data = filter_get_excel_data(file_path)
    for data in filter_data:
        print(data)
    search_data = search_get_excel_data(file_path)
    for data in search_data:
        print(data)